# -*- coding: utf-8 -*-
"""
影刀数据表格 - 理赔数据整理脚本
原理：数据分前后两段用同一套18列，前半段占A-E列，后半段占F-R列。
      将后半段数据向上平移到前半段行中，只在前端增加"城市"列，其余列原样保留。

v3: 增加花名册站点名称匹配功能
    合并后通过骑手姓名匹配花名册中的站点名称，同名按城市进一步筛选，
    城市也相同则按手机尾号后四位判断。
"""

import re
import math
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==================== 配置区域 ====================
INPUT_FILE = r"C:/Users/Administrator/Desktop/影刀数据表格.xlsx"
OUTPUT_FILE = r"C:/Users/Administrator/Desktop/影刀数据表格_整理后.xlsx"
ROSTER_FILE = r"C:/Users/Administrator/Desktop/骑手花名册.xlsx"
# =================================================

DATE_PATTERN = re.compile(r'^\d{4}-\d{2}-\d{2}')
CITY_NAMES = {'无锡','东莞','深圳','佛山','武汉','中山','西安','广州','惠州','南宁'}


def is_city_row(row):
    """单值行且是城市名 → 城市标记"""
    vals = row.dropna()
    return len(vals) == 1 and str(vals.iloc[0]) in CITY_NAMES


def is_front_half(row):
    """A列有数据（保单号），且不是城市标记、不是空行"""
    v0 = row.iloc[0] if pd.notna(row.iloc[0]) else ''
    return str(v0) and not is_city_row(row) and not DATE_PATTERN.match(str(v0))


def is_back_half(row):
    """A列为空，F列（出险时间）有数据"""
    a_empty = pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == ''
    f_val = row.iloc[5] if pd.notna(row.iloc[5]) else ''
    return a_empty and bool(DATE_PATTERN.match(str(f_val)))


def parse_amount(val):
    """金额 → ¥xxx.xx 或 '-'"""
    if val is None or pd.isna(val):
        return '-'
    s = str(val).strip()
    if s in ('-', '', 'nan', '0.00000'):
        return '-'
    try:
        num = float(s)
        if num == 0:
            return '-'
        return f"¥{num:,.2f}"
    except:
        return s


def extract_city_name(roster_city):
    """从花名册城市字段提取核心城市名
    例: '专送_东莞' → '东莞', 'UB_广州' → '广州',
        '城代_襄阳' → '襄阳', 'HY_城代_房县' → '房县'
    """
    if pd.isna(roster_city):
        return ''
    s = str(roster_city).strip()
    # 取最后一个下划线后的部分
    if '_' in s:
        return s.split('_')[-1]
    return s


def get_phone_tail4(phone_str):
    """提取手机号后4位（兼容脱敏格式如 '180****7823' 和完整号）"""
    if pd.isna(phone_str) or not phone_str:
        return ''
    s = str(phone_str).strip().replace('-', '').replace(' ', '')
    # 去掉所有非数字字符后取后4位
    digits = re.sub(r'\D', '', s)
    return digits[-4:] if len(digits) >= 4 else ''


def load_roster(roster_path):
    """加载花名册，构建姓名→记录列表的索引"""
    df_roster = pd.read_excel(roster_path, dtype={'站点ID': str, '电话': str})
    # 提取核心城市名
    df_roster['_核心城市'] = df_roster['城市'].apply(extract_city_name)
    # 提取电话后4位
    df_roster['_尾号4'] = df_roster['电话'].apply(get_phone_tail4)

    # 构建姓名索引
    name_index = {}
    for _, row in df_roster.iterrows():
        name = str(row['骑手姓名']).strip() if pd.notna(row['骑手姓名']) else ''
        if not name:
            continue
        if name not in name_index:
            name_index[name] = []
        name_index[name].append({
            '站点ID': str(row['站点ID']) if pd.notna(row['站点ID']) else '',
            '核心城市': row['_核心城市'],
            '尾号4': row['_尾号4'],
            '站点': str(row['站点']) if pd.notna(row['站点']) else '',
        })

    return name_index, df_roster


def match_site_name(name, city, phone, name_index):
    """三级匹配：姓名 → 城市 → 手机尾号后四位
    返回 (站点名称, 匹配方式)
    匹配方式: '唯一姓名' / '姓名+城市' / '姓名+城市+尾号' / '多重匹配' / '未匹配'
    """
    if not name or name == '-' or name == 'nan':
        return '', '未匹配'

    matches = name_index.get(name, [])
    if not matches:
        return '', '未匹配'

    # 第一级：姓名唯一匹配
    if len(matches) == 1:
        return matches[0]['站点'], '唯一姓名'

    # 第二级：按城市筛选
    city_matches = [m for m in matches if m['核心城市'] == city]
    if len(city_matches) == 1:
        return city_matches[0]['站点'], '姓名+城市'

    if len(city_matches) > 1:
        # 第三级：按手机尾号后4位筛选
        phone_tail = get_phone_tail4(phone)
        if phone_tail:
            phone_matches = [m for m in city_matches if m['尾号4'] == phone_tail]
            if len(phone_matches) == 1:
                return phone_matches[0]['站点'], '姓名+城市+尾号'
            elif len(phone_matches) > 1:
                # 仍然多匹配，返回所有站点名称用逗号拼接
                names = '/'.join(set(m['站点'] for m in phone_matches if m['站点']))
                return names, '多重匹配'

        # 城市匹配但尾号无法区分
        names = '/'.join(set(m['站点'] for m in city_matches if m['站点']))
        return names, '多重匹配(城市内)'

    # 姓名匹配但城市不匹配，尝试手机尾号兜底
    phone_tail = get_phone_tail4(phone)
    if phone_tail:
        phone_matches = [m for m in matches if m['尾号4'] == phone_tail]
        if len(phone_matches) == 1:
            return phone_matches[0]['站点'], '姓名+尾号(跨城市)'

    # 无法进一步区分
    names = '/'.join(set(m['站点'] for m in matches if m['站点']))
    return names, '多重匹配(跨城市)'


def process_data(df):
    """主流程：扫描原始数据，配对前后半段"""
    # 取表头（Row 0, col 0-17）
    raw_headers = [str(df.iloc[0, j]) if pd.notna(df.iloc[0, j]) else f'Col{j}' for j in range(18)]

    records = []
    current_city = None
    front_buf = []   # 前半段行号
    back_buf = []    # 后半段行号
    all_sections = []

    for i in range(1, len(df)):
        row = df.iloc[i]

        if is_city_row(row):
            # 保存上一个城市的数据
            if current_city is not None:
                all_sections.append((current_city, front_buf.copy(), back_buf.copy()))
            current_city = str(row.dropna().iloc[0])
            front_buf = []
            back_buf = []
        elif is_back_half(row):
            back_buf.append(i)
        elif is_front_half(row):
            if back_buf:
                # 已经进入后半段了但还是前半段格式 → 归入后半段延伸
                back_buf.append(i)
            else:
                front_buf.append(i)

    # 最后一个城市
    if current_city is not None:
        all_sections.append((current_city, front_buf.copy(), back_buf.copy()))

    print("\n城市分段解析:")
    for city, f_rows, b_rows in all_sections:
        match = 'OK' if len(f_rows) == len(b_rows) else f'!! 差{abs(len(f_rows)-len(b_rows))}'
        print(f"  {city}: 前半段{len(f_rows)}条, 后半段{len(b_rows)}条 {match}")

    # 合并：前半段全部输出为记录，有后半段数据的优先配对填充
    for city, f_rows, b_rows in all_sections:
        paired = min(len(f_rows), len(b_rows))
        extra = len(f_rows) - paired  # 仅基本信息的记录数

        for idx in range(len(f_rows)):
            f_row = df.iloc[f_rows[idx]]
            record = {'城市': city}
            # 前半段列: A-E = col 0-4
            for j in range(5):
                v = f_row.iloc[j]
                record[raw_headers[j]] = str(v) if pd.notna(v) else '-'
            # 后半段列: F-R = col 5-17
            if idx < len(b_rows):
                b_row = df.iloc[b_rows[idx]]
                for j in range(5, 18):
                    v = b_row.iloc[j]
                    col_name = raw_headers[j]
                    if col_name in ('赔款金额(元)', '估损总金额(元)'):
                        record[col_name] = parse_amount(v)
                    else:
                        record[col_name] = str(v).replace('● ', '') if pd.notna(v) else '-'
            else:
                for j in range(5, 18):
                    record[raw_headers[j]] = '-'

            records.append(record)

        if extra > 0:
            print(f"  {city}: {paired}条完整 + {extra}条仅基本信息")

    return records, raw_headers, all_sections


def match_records_with_roster(records, name_index):
    """为每条记录匹配花名册中的站点名称"""
    # raw_headers 中: col2=报案人, col3=骑报案人电话
    # records 中 key='报案人' 是姓名, key='骑报案人电话' 是电话
    match_stats = {
        '唯一姓名': 0, '姓名+城市': 0, '姓名+城市+尾号': 0,
        '姓名+尾号(跨城市)': 0, '多重匹配(城市内)': 0,
        '多重匹配(跨城市)': 0, '未匹配': 0,
    }

    for rec in records:
        name = rec.get('报案人', '-')
        city = rec.get('城市', '')
        phone = rec.get('骑报案人电话', '-')

        site_name, match_method = match_site_name(name, city, phone, name_index)
        rec['匹配站点名称'] = site_name if site_name else '-'
        rec['匹配方式'] = match_method
        match_stats[match_method] = match_stats.get(match_method, 0) + 1

    matched = sum(v for k, v in match_stats.items() if k != '未匹配')
    total = len(records)
    print(f"\n花名册匹配结果: {matched}/{total} 条成功匹配 ({matched/total*100:.1f}%)")
    for method, cnt in match_stats.items():
        if cnt > 0:
            print(f"  {method}: {cnt}条")

    return records, match_stats


def write_excel(records, raw_headers, all_sections):
    """写入格式化Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = '理赔数据汇总'

    # 最终表头: 城市 + 原18列 + 匹配站点名称 + 匹配方式
    final_headers = ['城市'] + raw_headers + ['匹配站点名称', '匹配方式']

    # === 样式 ===
    hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='微软雅黑', size=10)
    data_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    amt_font = Font(name='微软雅黑', size=10, color='C00000')
    zero_font = Font(name='微软雅黑', size=10, color='808080')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    status_colors = {
        '打款成功': 'E2EFDA',
        '待用户提交资料': 'FFF2CC',
        '注销': 'F2F2F2',
        '已撤案': 'F2F2F2',
        '0结': 'F2F2F2',
        '保险公司拒绝赔偿': 'FCE4D6',
    }

    # 匹配方式颜色标记
    match_colors = {
        '唯一姓名': 'E2EFDA',        # 浅绿 - 最可靠
        '姓名+城市': 'E2EFDA',       # 浅绿
        '姓名+城市+尾号': 'DAEEF3',   # 浅蓝 - 可靠
        '姓名+尾号(跨城市)': 'FFF2CC', # 浅黄 - 需关注
        '多重匹配(城市内)': 'FCE4D6',  # 浅红 - 不确定
        '多重匹配(跨城市)': 'FCE4D6',  # 浅红 - 不确定
    }

    # 写表头
    for ci, h in enumerate(final_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    # 写数据
    for ri, rec in enumerate(records, 2):
        for ci, h in enumerate(final_headers, 1):
            val = rec.get(h, '-')
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = data_font
            c.border = thin_border

            if h in ('赔款金额(元)', '估损总金额(元)'):
                if val == '-':
                    c.font = zero_font
                    c.alignment = data_align
                elif val.startswith('¥'):
                    c.font = amt_font
                    c.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    c.alignment = data_align
            elif h in ('保单号', '报案号', '关联报案号', '匹配站点名称'):
                c.alignment = left_align
            elif h == '匹配方式':
                c.alignment = data_align
                # 匹配方式颜色
                match_method = str(val)
                for kw, color in match_colors.items():
                    if kw in match_method:
                        c.fill = PatternFill('solid', fgColor=color)
                        break
                if match_method == '未匹配':
                    c.font = Font(name='微软雅黑', size=10, color='C00000')
            else:
                c.alignment = data_align

        # 案件状态行背景色
        status = rec.get('案件状态', '')
        for kw, color in status_colors.items():
            if kw in status:
                for cj in range(1, len(final_headers) + 1):
                    ws.cell(row=ri, column=cj).fill = PatternFill('solid', fgColor=color)
                break

    # === 列宽 ===
    col_count = len(final_headers)
    # 基础列宽 A-S (1-19)
    widths = [8, 34, 26, 10, 16, 20,
              20, 16, 14, 16, 16, 20,
              12, 14, 26, 12, 14, 14, 12]
    # 新增列宽: 匹配站点名称, 匹配方式
    widths += [20, 18]
    for ci, w in enumerate(widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(col_count)}{len(records)+1}'

    # === 汇总统计 ===
    ws2 = wb.create_sheet('汇总统计')
    stat_h = ['城市', '案件数', '已赔付金额(元)', '估损总金额(元)', '按时报案数', '投保人报案及时率',
              '匹配成功数', '匹配率']
    for ci, h in enumerate(stat_h, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = thin_border

    total_n = total_paid = total_est = total_timely = total_paired = 0
    total_matched = 0
    sr = 2
    for city, f_rows, b_rows in all_sections:
        n = len(f_rows)  # 全部案件数
        paired = min(n, len(b_rows))
        paid_sum = est_sum = timely_count = 0
        for idx in range(paired):
            f_row = df.iloc[f_rows[idx]]
            b_row = df.iloc[b_rows[idx]]
            try:
                p = float(b_row.iloc[8])
                if not math.isnan(p):
                    paid_sum += p
            except: pass
            try:
                e = float(b_row.iloc[9])
                if not math.isnan(e):
                    est_sum += e
            except: pass
            # 报案及时率: 报案时间(col4) - 出险时间(col5) <= 24小时
            try:
                t_report = datetime.strptime(str(f_row.iloc[4]).strip(), '%Y-%m-%d %H:%M:%S')
                t_accident = datetime.strptime(str(b_row.iloc[5]).strip(), '%Y-%m-%d %H:%M:%S')
                diff_h = (t_report - t_accident).total_seconds() / 3600
                if 0 <= diff_h <= 24:
                    timely_count += 1
            except: pass

        timely_rate = timely_count / paired * 100 if paired > 0 else 0

        # 该城市的匹配统计
        city_records = [r for r in records if r.get('城市') == city]
        city_matched = sum(1 for r in city_records if r.get('匹配站点名称', '-') != '-')
        match_rate = city_matched / len(city_records) * 100 if city_records else 0

        ws2.cell(row=sr, column=1, value=city).font = Font(name='微软雅黑', bold=True, size=10)
        ws2.cell(row=sr, column=2, value=n).font = data_font
        ws2.cell(row=sr, column=3, value=paid_sum).font = data_font
        ws2.cell(row=sr, column=3).number_format = '¥#,##0.00'
        ws2.cell(row=sr, column=4, value=est_sum).font = data_font
        ws2.cell(row=sr, column=4).number_format = '¥#,##0.00'
        ws2.cell(row=sr, column=5, value=timely_count).font = data_font
        ws2.cell(row=sr, column=6, value=f'{timely_rate:.1f}%').font = data_font
        if timely_rate >= 80:
            ws2.cell(row=sr, column=6).font = Font(name='微软雅黑', size=10, color='008000', bold=True)
        elif timely_rate < 50:
            ws2.cell(row=sr, column=6).font = Font(name='微软雅黑', size=10, color='C00000', bold=True)
        ws2.cell(row=sr, column=7, value=city_matched).font = data_font
        ws2.cell(row=sr, column=8, value=f'{match_rate:.1f}%').font = data_font
        if match_rate >= 80:
            ws2.cell(row=sr, column=8).font = Font(name='微软雅黑', size=10, color='008000', bold=True)
        elif match_rate < 50:
            ws2.cell(row=sr, column=8).font = Font(name='微软雅黑', size=10, color='C00000', bold=True)
        for cj in range(1, 9):
            ws2.cell(row=sr, column=cj).alignment = data_align
            ws2.cell(row=sr, column=cj).border = thin_border
        if paid_sum > 0:
            ws2.cell(row=sr, column=3).font = Font(name='微软雅黑', size=10, color='C00000')
        total_n += n; total_paid += paid_sum; total_est += est_sum; total_timely += timely_count; total_paired += paired
        total_matched += city_matched
        sr += 1

    # 合计行
    total_rate = total_timely / total_paired * 100 if total_paired > 0 else 0
    total_match_rate = total_matched / len(records) * 100 if records else 0
    ws2.cell(row=sr, column=1, value='合计').font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    ws2.cell(row=sr, column=1).fill = PatternFill('solid', fgColor='2F5496')
    ws2.cell(row=sr, column=2, value=total_n).font = Font(name='微软雅黑', bold=True, size=11)
    ws2.cell(row=sr, column=3, value=total_paid).font = Font(name='微软雅黑', bold=True, size=11, color='C00000')
    ws2.cell(row=sr, column=3).number_format = '¥#,##0.00'
    ws2.cell(row=sr, column=4, value=total_est).font = Font(name='微软雅黑', bold=True, size=11)
    ws2.cell(row=sr, column=4).number_format = '¥#,##0.00'
    ws2.cell(row=sr, column=5, value=total_timely).font = Font(name='微软雅黑', bold=True, size=11)
    ws2.cell(row=sr, column=6, value=f'{total_rate:.1f}%').font = Font(name='微软雅黑', bold=True, size=11, color='2F5496')
    ws2.cell(row=sr, column=7, value=total_matched).font = Font(name='微软雅黑', bold=True, size=11)
    ws2.cell(row=sr, column=8, value=f'{total_match_rate:.1f}%').font = Font(name='微软雅黑', bold=True, size=11, color='2F5496')
    for cj in range(1, 9):
        ws2.cell(row=sr, column=cj).alignment = data_align
        ws2.cell(row=sr, column=cj).border = Border(
            left=Side(style='thin', color='2F5496'), right=Side(style='thin', color='2F5496'),
            top=Side(style='medium', color='2F5496'), bottom=Side(style='medium', color='2F5496'))
        ws2.cell(row=sr, column=cj).fill = PatternFill('solid', fgColor='D6E4F0')

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 18
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 10
    ws2.freeze_panes = 'A2'

    # === 匹配明细工作表 ===
    ws3 = wb.create_sheet('匹配明细')
    detail_h = ['城市', '报案人', '骑报案人电话', '匹配站点名称', '匹配方式', '花名册城市', '花名册站点', '花名册电话']
    for ci, h in enumerate(detail_h, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = thin_border

    # 重新读取花名册以获取详细信息用于匹配明细
    # 只输出匹配成功或需要关注的记录
    dr = 2
    for rec in records:
        match_method = rec.get('匹配方式', '未匹配')
        if match_method == '未匹配':
            continue  # 跳过未匹配的
        c = ws3.cell(row=dr, column=1, value=rec.get('城市', ''))
        c.font = data_font; c.border = thin_border; c.alignment = data_align
        c = ws3.cell(row=dr, column=2, value=rec.get('报案人', '-'))
        c.font = data_font; c.border = thin_border; c.alignment = data_align
        c = ws3.cell(row=dr, column=3, value=rec.get('骑报案人电话', '-'))
        c.font = data_font; c.border = thin_border; c.alignment = data_align
        c = ws3.cell(row=dr, column=4, value=rec.get('匹配站点名称', '-'))
        c.font = data_font; c.border = thin_border; c.alignment = data_align
        c = ws3.cell(row=dr, column=5, value=match_method)
        c.font = data_font; c.border = thin_border; c.alignment = data_align
        # 匹配方式颜色
        for kw, color in match_colors.items():
            if kw in match_method:
                for cj in range(1, 9):
                    ws3.cell(row=dr, column=cj).fill = PatternFill('solid', fgColor=color)
                break
        dr += 1

    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 16
    ws3.column_dimensions['E'].width = 20
    ws3.column_dimensions['F'].width = 14
    ws3.column_dimensions['G'].width = 24
    ws3.column_dimensions['H'].width = 16
    ws3.freeze_panes = 'A2'

    wb.save(OUTPUT_FILE)
    print(f"\n输出文件: {OUTPUT_FILE}")


def main():
    print("=" * 60)
    print("影刀数据表格 - 理赔数据整理 (v3 - 含花名册站点名称匹配)")
    print("=" * 60)

    global df
    df = pd.read_excel(INPUT_FILE, header=None)
    print(f"\n读取: {INPUT_FILE}")
    print(f"总行数: {len(df)}, 总列数: {len(df.columns)}")

    records, raw_headers, all_sections = process_data(df)
    print(f"\n合并完成, 共 {len(records)} 条记录")

    # 加载花名册并匹配站点名称
    print(f"\n⏳ 正在加载花名册: {ROSTER_FILE}")
    name_index, df_roster = load_roster(ROSTER_FILE)
    print(f"   花名册总人数: {len(df_roster)}, 唯一姓名数: {len(name_index)}")

    print("⏳ 正在匹配站点名称...")
    records, match_stats = match_records_with_roster(records, name_index)

    write_excel(records, raw_headers, all_sections)

    # 状态分布
    status_count = {}
    for r in records:
        s = r.get('案件状态', '-')
        status_count[s] = status_count.get(s, 0) + 1
    print("\n案件状态分布:")
    for s, c in sorted(status_count.items(), key=lambda x: -x[1]):
        print(f"  {s}: {c}件")
    print("=" * 60)


if __name__ == '__main__':
    main()
