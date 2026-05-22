# -*- coding: utf-8 -*-
"""
考试酷(examcoo.com) 成绩批量下载脚本
- 自动获取pid、班级名称、考试名称
- 分页获取全部成绩数据
- 保留每题得分，每个学员只保留最高分
- 不同班级的成绩保存到不同的Excel文件，每场考试一个Sheet
"""

import requests
import re
import json
import os
import time
import urllib3
import pandas as pd

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# 禁用SSL警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ===================== 配置区域 =====================

# Cookie（从浏览器开发者工具复制，完整Cookie字符串）
COOKIE = "lastspl=dae698870e6af9d5deb518acf97bf32a; lastuid=15298143; autoUid=15298143; autoPassword=39b5eef9ec33a632bdce101960533acf; autoTimezone=28800; PHPSESSID=07ivfei3lr36tsjbl0skt044f4"

# 考试列表：每项填 cid(班级ID) 和 tid(考试ID)
# 从URL中提取：/class/score/paperanalysis/cid/{cid}/tid/{tid}/idpaging/0
EXAMS = [
    #头狼考试

    {"cid": "1100056", "tid": "5628408"},#第1节
    {"cid": "1100056", "tid": "5639097"},#第2节
    {"cid": "1100056", "tid": "5654168"},#第3节
    {"cid": "1100056", "tid": "5654167"},#第3-1节
    {"cid": "1100056", "tid": "5666173"},#第4节
    {"cid": "1100056", "tid": "5676874"},#第5节
    {"cid": "1100056", "tid": "5676875"},#第6节
    {"cid": "1100056", "tid": "5685392"},#第7节
    {"cid": "1100056", "tid": "5697851"},#第8节
    {"cid": "1100056", "tid": "5700569"},#第9节 
    {"cid": "1100056", "tid": "5700573"},#第10节


    #树人考试
    {"cid": "1099008", "tid": "5628406"},#第1节
    {"cid": "1099008", "tid": "5639096"},#第2节
    {"cid": "1099008", "tid": "5654169"},#第3节
    {"cid": "1099008", "tid": "5666172"},#第4节
    {"cid": "1099008", "tid": "5676872"},#第5节
    {"cid": "1099008", "tid": "5676873"},#第6节
    {"cid": "1099008", "tid": "5685391"},#第7节
    {"cid": "1099008", "tid": "5697850"},#第8节
    {"cid": "1099008", "tid": "5700559"},#第9节
    {"cid": "1099008", "tid": "5700566"},#第10节
]

# 输出目录（默认桌面）
OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "Desktop")

# 请求间隔（秒），避免被限流
REQUEST_DELAY = 0.5

# 最大分页步数（安全限制）
MAX_STEPS = 30

# ===================== 核心逻辑 =====================

BASE_URL = "https://www.examcoo.com"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36 Edg/148.0.0.0",
    "Cookie": COOKIE,
}

# 题型映射
TYPE_MAP = {
    "1": "单选",
    "2": "多选",
    "3": "判断",
    "4": "填空",
    "5": "主观",
}


def get_page_info(cid, tid):
    """
    从试卷分析HTML页面提取 pid、班级名称、考试名称
    """
    url = f"{BASE_URL}/class/score/paperanalysis/cid/{cid}/tid/{tid}/idpaging/0"
    resp = requests.get(url, headers=HEADERS, verify=False, timeout=30)
    html = resp.text

    # 提取 gPid
    pid = None
    m = re.search(r'var\s+gPid\s*=\s*"(\d+)"', html)
    if m:
        pid = m.group(1)

    # 提取班级名称（格式：班级名称／cid）
    # HTML中用 &#12288; 表示全角空格，用更灵活的匹配
    class_name = ""
    m = re.search(r'班.*?级[：:].*?<td>(.*?)(?:／|<)', html, re.DOTALL)
    if m:
        class_name = m.group(1).strip()

    # 提取考试名称（格式：考试名称／pid）
    exam_name = ""
    m = re.search(r'考试试卷[：:].*?<td>(.*?)(?:／|<)', html, re.DOTALL)
    if m:
        exam_name = m.group(1).strip()

    return pid, class_name, exam_name


def fetch_all_exam_data(cid, tid, pid):
    """
    分页获取全部考试数据
    - step=0 返回 paperData（题目结构）+ 首批 examerData
    - hasData=True 时继续获取下一页
    """
    all_paper_data = None
    all_examer_data = []
    step = 0

    while step < MAX_STEPS:
        post_data = {
            "pid": pid,
            "tid": tid,
            "cid": cid,
            "idpaging": "0",
            "step": str(step),
        }
        referer = f"{BASE_URL}/class/score/paperanalysis/cid/{cid}/tid/{tid}/idpaging/0"
        req_headers = {**HEADERS, "Referer": referer}

        try:
            resp = requests.post(
                f"{BASE_URL}/class/score/examcontent",
                headers=req_headers,
                data=post_data,
                verify=False,
                timeout=60,
            )
            result = resp.json()
        except Exception as e:
            print(f"  [错误] 第{step}页请求失败: {e}")
            break

        # 检查返回值是否为对象
        if not isinstance(result, dict):
            print(f"  [错误] 第{step}页返回异常: {result}")
            break

        # step=0 时保存 paperData
        if step == 0:
            all_paper_data = result.get("paperData", [])

        examer_data = result.get("examerData", [])
        all_examer_data.extend(examer_data)
        print(f"  第{step}页: 获取 {len(examer_data)} 条记录")

        # 检查是否还有更多数据
        has_data = result.get("hasData")
        if not has_data or has_data == 0 or has_data == "0" or has_data is False:
            break

        step += 1
        time.sleep(REQUEST_DELAY)

    return all_paper_data, all_examer_data


def parse_questions(paper_data):
    """
    解析试卷结构，返回题目列表
    - id 以 's' 开头的是具体题目
    - id 以 'b' 开头的是大题分组（跳过）
    """
    questions = []
    q_num = 0
    for item in paper_data:
        qid = item.get("id", "")
        if not qid.startswith("s"):
            continue
        q_num += 1
        type_code = qid.split("_")[0][1:]  # 提取 s 后面的数字
        q_type = TYPE_MAP.get(type_code, f"题型{type_code}")
        questions.append({
            "id": qid,
            "num": q_num,
            "type": q_type,
            "points": float(item.get("p", 0)),
        })
    return questions


def parse_student_scores(examer_data, questions):
    """
    解析学员成绩数据
    - 提取每题得分（从 mark_content 中获取 s 字段）
    - 同一学员多次考试只保留最高总分
    """
    # 题目 id -> 索引映射
    q_id_set = {q["id"] for q in questions}

    students = {}  # key: uid, value: {uid, name, totalscore, q_scores}

    for item in examer_data:
        uid = str(item.get("uid", ""))
        if not uid:
            continue

        # 确定学员姓名
        name = item.get("userremark", "")
        if not name or name == "None":
            name = item.get("rsstudentname", "")
        if not name or name == "None":
            name = item.get("displayname", "")
        if not name or name == "None":
           
            name = f"用户{uid}"

        # 学号（如果有的话）
        student_id = item.get("rsstudentid", "")
        if not student_id or student_id == "None":
            student_id = uid

        total = float(item.get("totalscore", 0))

        # 解析每题得分
        mark_content_str = item.get("mark_content", "[]")
        try:
            mark_content = json.loads(mark_content_str)
        except (json.JSONDecodeError, TypeError):
            mark_content = []

        q_scores = {}
        for mq in mark_content:
            qid = mq.get("id", "")
            if qid not in q_id_set:
                continue
            score_val = mq.get("s", "0")
            # 某些题型可能有逗号分隔的多部分得分
            if isinstance(score_val, str) and "," in score_val:
                try:
                    score_val = sum(float(p) for p in score_val.split(","))
                except ValueError:
                    score_val = 0.0
            else:
                try:
                    score_val = float(score_val)
                except (ValueError, TypeError):
                    score_val = 0.0
            q_scores[qid] = score_val

        # 只保留最高分
        if uid in students:
            if total > students[uid]["totalscore"]:
                students[uid] = {
                    "uid": uid,
                    "student_id": student_id,
                    "name": name,
                    "totalscore": total,
                    "q_scores": q_scores,
                }
        else:
            students[uid] = {
                "uid": uid,
                "student_id": student_id,
                "name": name,
                "totalscore": total,
                "q_scores": q_scores,
            }

    return list(students.values())


def build_dataframe(students, questions):
    """
    构建结果 DataFrame
    列：学号/账号、姓名、总分、第1题(题型,分值)、第2题(题型,分值)...
    """
    rows = []
    for s in students:
        row = {
            "学号/账号": s["student_id"],
            "姓名": s["name"],
            "总分": s["totalscore"],
        }
        for q in questions:
            col_name = f"第{q['num']}题({q['type']},{q['points']}分)"
            row[col_name] = s["q_scores"].get(q["id"], 0)
        rows.append(row)

    df = pd.DataFrame(rows)
    df = df.sort_values("总分", ascending=False).reset_index(drop=True)
    return df


def safe_filename(name):
    """将名称转为安全的文件名"""
    return re.sub(r'[\\/:*?"<>|]', '_', name)


def safe_sheet_name(name):
    """将名称转为安全的Sheet名（最长31字符）"""
    safe = re.sub(r'[\\/:*?"<>|]', '_', name)
    return safe[:31]


def process_exam(cid, tid):
    """处理单个考试，返回 (class_name, exam_name, DataFrame)"""
    print(f"\n{'='*60}")
    print(f"正在处理: cid={cid}, tid={tid}")

    # 1. 获取页面信息
    pid, class_name, exam_name = get_page_info(cid, tid)
    if not pid:
        print(f"  [错误] 无法获取pid，请检查Cookie是否有效")
        return None
    print(f"  班级: {class_name}")
    print(f"  考试: {exam_name}")
    print(f"  pid: {pid}")

    # 2. 分页获取数据
    paper_data, examer_data = fetch_all_exam_data(cid, tid, pid)
    print(f"  题目数: {len(paper_data) if paper_data else 0}")
    print(f"  考试人次: {len(examer_data)}")

    if not paper_data or not examer_data:
        print(f"  [警告] 无数据，跳过")
        return None

    # 3. 解析题目结构
    questions = parse_questions(paper_data)
    print(f"  有效题目: {len(questions)}")

    # 4. 解析学员成绩（去重保留最高分）
    students = parse_student_scores(examer_data, questions)
    print(f"  独立学员数(去重保留最高分): {len(students)}")

    # 5. 构建 DataFrame
    df = build_dataframe(students, questions)
    print(f"  DataFrame 行数: {len(df)}, 列数: {len(df.columns)}")

    return class_name, exam_name, df


def main():
    """主函数"""
    print("=" * 60)
    print("考试酷成绩批量下载工具")
    print("=" * 60)

    if not EXAMS:
        print("[错误] 未配置考试列表，请在脚本顶部 EXAMS 中添加考试信息")
        return

    # 按班级分组收集结果
    class_data = {}  # class_name -> [(exam_name, df), ...]

    for exam in EXAMS:
        cid = exam["cid"]
        tid = exam["tid"]
        result = process_exam(cid, tid)
        if result:
            class_name, exam_name, df = result
            if class_name not in class_data:
                class_data[class_name] = []
            class_data[class_name].append((exam_name, df))

    # 保存到Excel - 不同班级不同文件，同一班级不同考试不同Sheet
    if not class_data:
        print("\n[错误] 没有获取到任何数据")
        return

    saved_files = []
    for class_name, exams in class_data.items():
        safe_name = safe_filename(class_name)
        filepath = os.path.join(OUTPUT_DIR, f"{safe_name}_成绩.xlsx")

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            for exam_name, df in exams:
                sheet = safe_sheet_name(exam_name)
                df.to_excel(writer, sheet_name=sheet, index=False)

                # 自动调整列宽
                worksheet = writer.sheets[sheet]
                for col_idx, col in enumerate(df.columns, 1):
                    # 计算最大宽度
                    max_len = max(
                        df[col].astype(str).map(len).max(),
                        len(str(col))
                    )
                    # 中文字符按2个宽度计算
                    col_len = 0
                    for char in str(col):
                        col_len += 2 if ord(char) > 127 else 1
                    for val in df[col].astype(str):
                        val_len = 0
                        for char in val:
                            val_len += 2 if ord(char) > 127 else 1
                        col_len = max(col_len, val_len)
                    adjusted_width = min(col_len + 4, 50)
                    worksheet.column_dimensions[
                        worksheet.cell(row=1, column=col_idx).column_letter
                    ].width = adjusted_width

        saved_files.append(filepath)
        print(f"\n已保存: {filepath}")

    print(f"\n{'='*60}")
    print(f"全部完成！共保存 {len(saved_files)} 个Excel文件:")
    for f in saved_files:
        print(f"  📁 {f}")


if __name__ == "__main__":
    main()




# ---- 1. 读取数据 ----
roster = pd.read_excel('C:/Users/Administrator/Desktop/树人计划第六期/学员名单汇总.xlsx', sheet_name='Sheet1')
xl_sren = pd.read_excel('C:/Users/Administrator/Desktop/树人计划第6期训练营_成绩.xlsx', sheet_name=None)
xl_toul = pd.read_excel('C:/Users/Administrator/Desktop/头狼计划第6期训练营_成绩.xlsx', sheet_name=None)

# ---- 2. 姓名清洗函数：提取末尾的真实姓名 ----
# 例如 "洪梅站-闲宇航" -> "闲宇航"，"广州壹城-中山-沙朗站李汉权" -> "李汉权"
# 规则：如果含 "-"，取最后一个 "-" 之后的部分；否则原样
def clean_name(name):
    name = str(name).strip()
    if '-' in name:
        return name.split('-')[-1].strip()
    return name

# 对成绩文件所有sheet做姓名清洗
def clean_sheet_names(sheets_dict):
    cleaned = {}
    for sheet_name, df in sheets_dict.items():
        df = df.copy()
        if '姓名' in df.columns:
            df['姓名'] = df['姓名'].apply(clean_name)
        cleaned[sheet_name] = df
    return cleaned

xl_sren = clean_sheet_names(xl_sren)
xl_toul = clean_sheet_names(xl_toul)

# ---- 3. 定义课程列表 ----
# 树人课程
sren_courses = list(xl_sren.keys())
# 头狼课程
toul_courses = list(xl_toul.keys())

print('树人课程:', sren_courses)
print('头狼课程:', toul_courses)

# ---- 4. 构建每个课程、每个学员的最高折算分 ----
def get_max_score_per_person(sheets_dict):
    """
    返回 dict: {sheet_name: {姓名: 折算后最高分}}
    折算规则：若该sheet最高总分 > 100，则折算为 (总分/该sheet最高总分) * 100 取整
    若总分 <= 100，直接使用
    """
    result = {}
    for sheet_name, df in sheets_dict.items():
        if '姓名' not in df.columns or '总分' not in df.columns:
            continue
        df = df.copy()
        df['总分'] = pd.to_numeric(df['总分'], errors='coerce')
        
        max_possible = df['总分'].max()
        
        # 按姓名取最高分
        best = df.groupby('姓名')['总分'].max().reset_index()
        
        name_score = {}
        for _, row in best.iterrows():
            raw_score = row['总分']
            if pd.isna(raw_score):
                continue
            if max_possible > 100:
                # 折算成100分制，保留1位小数
                converted = round(raw_score / max_possible * 100, 1)
            else:
                converted = raw_score
            name_score[row['姓名']] = converted
        
        result[sheet_name] = name_score
    return result

sren_scores = get_max_score_per_person(xl_sren)
toul_scores = get_max_score_per_person(xl_toul)

# ---- 5. 判断归属群，区分树人学员 vs 头狼学员 ----
# 树人群: 树人1群
# 头狼群: 头狼1群, 头狼2群
roster['_is_sren'] = roster['归属群'].str.contains('树人', na=False)
roster['_is_toul'] = roster['归属群'].str.contains('头狼', na=False)

# ---- 6. 生成汇总数据 ----
# 根据归属群选用不同的成绩课程
# 树人学员用树人成绩课程，头狼学员用头狼成绩课程
# 但两类学员都在同一名单，需要将对应的列补全

# 将树人课程和头狼课程列名稍作说明（防止重名）
# 检查是否有重名的sheet
sren_set = set(sren_courses)
toul_set = set(toul_courses)
overlap = sren_set & toul_set
print('重名课程:', overlap)

# 重名的课程加前缀区分
def prefix_courses(courses, prefix):
    renamed = {}
    for c in courses:
        if c in overlap:
            renamed[c] = f'[{prefix}]{c}'
        else:
            renamed[c] = c
    return renamed

sren_renamed = prefix_courses(sren_courses, '树人')
toul_renamed = prefix_courses(toul_courses, '头狼')

# ---- 7. 构建输出 DataFrame ----
# 基础列：归属群、序号、姓名、城市、站点
# 然后跟树人各节成绩、头狼各节成绩（每人只显示自己归属的成绩，另一类的成绩留空）

all_cols = ['归属群', '序号', '姓名', '城市', '站点']
# 树人课程列
sren_col_names = [sren_renamed[c] for c in sren_courses]
# 头狼课程列
toul_col_names = [toul_renamed[c] for c in toul_courses]

rows = []
for _, r in roster.iterrows():
    name = r['姓名']
    row = {
        '归属群': r['归属群'],
        '序号': r['序号'],
        '姓名': name,
        '城市': r['城市'],
        '站点': r['站点'],
    }
    
    # 根据归属群填写成绩
    is_sren = r['_is_sren']
    is_toul = r['_is_toul']
    
    # 树人课程成绩
    for c in sren_courses:
        col = sren_renamed[c]
        if is_sren:
            score_map = sren_scores.get(c, {})
            if name in score_map:
                row[col] = score_map[name]
            else:
                row[col] = '未完成'
        else:
            row[col] = ''  # 头狼学员不参加树人课程，留空
    
    # 头狼课程成绩
    for c in toul_courses:
        col = toul_renamed[c]
        if is_toul:
            score_map = toul_scores.get(c, {})
            if name in score_map:
                row[col] = score_map[name]
            else:
                row[col] = '未完成'
        else:
            row[col] = ''  # 树人学员不参加头狼课程，留空
    
    rows.append(row)

df_out = pd.DataFrame(rows)

# 计算树人学员平均分和头狼学员平均分
def calc_avg(row, course_cols, is_flag):
    """计算有效数字成绩的平均分"""
    if not is_flag:
        return ''
    scores = []
    for c in course_cols:
        v = row.get(c, '')
        if isinstance(v, (int, float)) and not pd.isna(v):
            scores.append(v)
        # 字符串'未完成'不计入
    if scores:
        return round(sum(scores) / len(scores), 1)
    return ''

sren_col_vals = sren_col_names
toul_col_vals = toul_col_names

# 添加完成率列
def calc_completion(row, course_cols, is_flag):
    if not is_flag:
        return ''
    completed = 0
    for c in course_cols:
        v = row.get(c, '')
        if isinstance(v, (int, float)) and not pd.isna(v):
            completed += 1
    total = len(course_cols)
    if total == 0:
        return ''
    return f'{completed}/{total}'

# 暂不加平均分列，直接输出原始成绩

print('输出列:', df_out.columns.tolist())
print(f'总行数: {len(df_out)}')

# ---- 8. 写出 Excel ----
output_path = 'C:/Users/Administrator/Desktop/树人计划第六期/学员成绩汇总.xlsx'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '成绩汇总'

# 定义样式
header_font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
header_fill_sren = PatternFill('solid', start_color='2E75B6')   # 蓝色-树人
header_fill_toul = PatternFill('solid', start_color='375623')   # 深绿-头狼
header_fill_base = PatternFill('solid', start_color='404040')   # 深灰-基础信息
cell_font = Font(name='微软雅黑', size=9)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 写表头
header_row = ['归属群', '序号', '姓名', '城市', '站点'] + sren_col_names + toul_col_names
for ci, h in enumerate(header_row, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border
    if ci <= 5:
        cell.fill = header_fill_base
    elif ci <= 5 + len(sren_col_names):
        cell.fill = header_fill_sren
    else:
        cell.fill = header_fill_toul

# 写数据
sren_member_fill = PatternFill('solid', start_color='DDEEFF')  # 浅蓝-树人成员行
toul_member_fill = PatternFill('solid', start_color='E8F0E0')  # 浅绿-头狼成员行
unfinished_fill = PatternFill('solid', start_color='FFE0E0')   # 浅红-未完成
score_font = Font(name='微软雅黑', size=9)
unfinished_font = Font(name='微软雅黑', size=9, color='CC0000')

all_col_keys = ['归属群', '序号', '姓名', '城市', '站点'] + sren_col_names + toul_col_names

# 按归属群排序：树人1群，头狼1群，头狼2群
df_out['_sort'] = df_out['归属群'].map({'树人1群': 0, '头狼1群': 1, '头狼2群': 2}).fillna(3)
df_out = df_out.sort_values(['_sort', '序号']).reset_index(drop=True)

for ri, row in df_out.iterrows():
    excel_row = ri + 2
    is_sren = '树人' in str(row.get('归属群', ''))
    row_bg = sren_member_fill if is_sren else toul_member_fill
    
    for ci, col in enumerate(all_col_keys, 1):
        val = row.get(col, '')
        cell = ws.cell(row=excel_row, column=ci, value=val)
        cell.font = score_font
        cell.alignment = center
        cell.border = border
        
        # 未完成单元格特殊标色
        if val == '未完成':
            cell.fill = unfinished_fill
            cell.font = unfinished_font
        else:
            cell.fill = row_bg

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 6
ws.column_dimensions['E'].width = 10
# 课程列宽度
for i in range(6, len(header_row) + 1):
    ws.column_dimensions[get_column_letter(i)].width = 14

# 冻结前5列和首行
ws.freeze_panes = 'F2'

# 第一行行高
ws.row_dimensions[1].height = 40

# ---- 9. 添加第二个sheet：说明 ----
ws2 = wb.create_sheet('折算说明')
ws2['A1'] = '折算规则说明'
ws2['A1'].font = Font(bold=True, size=12)
notes = [
    ['', ''],
    ['规则', '说明'],
    ['最高分保留', '同一学员在同一节课有多条成绩记录时，保留最高分'],
    ['折算100分制', '若该节课总分上限超过100分，则按比例折算：折算分 = 原始分 / 满分 * 100（保留1位小数）'],
    ['未完成', '学员在名单中但该节课无成绩记录，显示"未完成"'],
    ['空白', '该节课不属于该学员所在群组，不做统计'],
    ['', ''],
    ['树人成绩说明', ''],
]
for c in sren_courses:
    df = xl_sren[c] if c in xl_sren else None
    if df is not None and '总分' in df.columns:
        max_s = pd.to_numeric(df['总分'], errors='coerce').max()
        notes.append([c, f'满分={max_s}分' + ('（需折算）' if max_s > 100 else '（无需折算）')])
notes.append(['', ''])
notes.append(['头狼成绩说明', ''])
for c in toul_courses:
    df = xl_toul[c] if c in xl_toul else None
    if df is not None and '总分' in df.columns:
        max_s = pd.to_numeric(df['总分'], errors='coerce').max()
        col_label = toul_renamed[c]
        notes.append([col_label, f'满分={max_s}分' + ('（需折算）' if max_s > 100 else '（无需折算）')])

for r_data in notes:
    ws2.append(r_data)

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 30

wb.save(output_path)
print(f'\n已保存至: {output_path}')

# 打印匹配统计
print('\n=== 匹配统计 ===')
for c in sren_courses:
    score_map = sren_scores.get(c, {})
    matched = sum(1 for _, r in roster[roster['_is_sren']].iterrows() if r['姓名'] in score_map)
    total_sren = roster['_is_sren'].sum()
    print(f'[树人]{c}: {matched}/{total_sren} 人有成绩')

for c in toul_courses:
    score_map = toul_scores.get(c, {})
    matched = sum(1 for _, r in roster[roster['_is_toul']].iterrows() if r['姓名'] in score_map)
    total_toul = roster['_is_toul'].sum()
    print(f'[头狼]{c}: {matched}/{total_toul} 人有成绩')


import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# ---- 1. 读取数据 ----
roster = pd.read_excel('C:/Users/Administrator/Desktop/树人计划第六期/学员名单汇总.xlsx', sheet_name='Sheet1')
xl_sren = pd.read_excel('C:/Users/Administrator/Desktop/树人计划第6期训练营_成绩.xlsx', sheet_name=None)
xl_toul = pd.read_excel('C:/Users/Administrator/Desktop/头狼计划第6期训练营_成绩.xlsx', sheet_name=None)

# ---- 2. 姓名清洗函数：提取末尾的真实姓名 ----
# 例如 "洪梅站-闲宇航" -> "闲宇航"，"广州壹城-中山-沙朗站李汉权" -> "李汉权"
# 规则：如果含 "-"，取最后一个 "-" 之后的部分；否则原样
def clean_name(name):
    name = str(name).strip()
    if '-' in name:
        return name.split('-')[-1].strip()
    return name

# 对成绩文件所有sheet做姓名清洗
def clean_sheet_names(sheets_dict):
    cleaned = {}
    for sheet_name, df in sheets_dict.items():
        df = df.copy()
        if '姓名' in df.columns:
            df['姓名'] = df['姓名'].apply(clean_name)
        cleaned[sheet_name] = df
    return cleaned

xl_sren = clean_sheet_names(xl_sren)
xl_toul = clean_sheet_names(xl_toul)

# ---- 3. 定义课程列表 ----
# 树人课程
sren_courses = list(xl_sren.keys())
# 头狼课程
toul_courses = list(xl_toul.keys())

print('树人课程:', sren_courses)
print('头狼课程:', toul_courses)

# ---- 4. 构建每个课程、每个学员的最高折算分 ----
def get_max_score_per_person(sheets_dict):
    """
    返回 dict: {sheet_name: {姓名: 折算后最高分}}
    折算规则：若该sheet最高总分 > 100，则折算为 (总分/该sheet最高总分) * 100 取整
    若总分 <= 100，直接使用
    """
    result = {}
    for sheet_name, df in sheets_dict.items():
        if '姓名' not in df.columns or '总分' not in df.columns:
            continue
        df = df.copy()
        df['总分'] = pd.to_numeric(df['总分'], errors='coerce')
        
        max_possible = df['总分'].max()
        
        # 按姓名取最高分
        best = df.groupby('姓名')['总分'].max().reset_index()
        
        name_score = {}
        for _, row in best.iterrows():
            raw_score = row['总分']
            if pd.isna(raw_score):
                continue
            if max_possible > 100:
                # 折算成100分制，保留1位小数
                converted = round(raw_score / max_possible * 100, 1)
            else:
                converted = raw_score
            name_score[row['姓名']] = converted
        
        result[sheet_name] = name_score
    return result

sren_scores = get_max_score_per_person(xl_sren)
toul_scores = get_max_score_per_person(xl_toul)

# ---- 5. 判断归属群，区分树人学员 vs 头狼学员 ----
# 树人群: 树人1群
# 头狼群: 头狼1群, 头狼2群
roster['_is_sren'] = roster['归属群'].str.contains('树人', na=False)
roster['_is_toul'] = roster['归属群'].str.contains('头狼', na=False)

# ---- 6. 生成汇总数据 ----
# 根据归属群选用不同的成绩课程
# 树人学员用树人成绩课程，头狼学员用头狼成绩课程
# 但两类学员都在同一名单，需要将对应的列补全

# 将树人课程和头狼课程列名稍作说明（防止重名）
# 检查是否有重名的sheet
sren_set = set(sren_courses)
toul_set = set(toul_courses)
overlap = sren_set & toul_set
print('重名课程:', overlap)

# 重名的课程加前缀区分
def prefix_courses(courses, prefix):
    renamed = {}
    for c in courses:
        if c in overlap:
            renamed[c] = f'[{prefix}]{c}'
        else:
            renamed[c] = c
    return renamed

sren_renamed = prefix_courses(sren_courses, '树人')
toul_renamed = prefix_courses(toul_courses, '头狼')

# ---- 7. 构建输出 DataFrame ----
# 基础列：归属群、序号、姓名、城市、站点
# 然后跟树人各节成绩、头狼各节成绩（每人只显示自己归属的成绩，另一类的成绩留空）

all_cols = ['归属群', '序号', '姓名', '城市', '站点']
# 树人课程列
sren_col_names = [sren_renamed[c] for c in sren_courses]
# 头狼课程列
toul_col_names = [toul_renamed[c] for c in toul_courses]

rows = []
for _, r in roster.iterrows():
    name = r['姓名']
    row = {
        '归属群': r['归属群'],
        '序号': r['序号'],
        '姓名': name,
        '城市': r['城市'],
        '站点': r['站点'],
    }
    
    # 根据归属群填写成绩
    is_sren = r['_is_sren']
    is_toul = r['_is_toul']
    
    # 树人课程成绩
    for c in sren_courses:
        col = sren_renamed[c]
        if is_sren:
            score_map = sren_scores.get(c, {})
            if name in score_map:
                row[col] = score_map[name]
            else:
                row[col] = '未完成'
        else:
            row[col] = ''  # 头狼学员不参加树人课程，留空
    
    # 头狼课程成绩
    for c in toul_courses:
        col = toul_renamed[c]
        if is_toul:
            score_map = toul_scores.get(c, {})
            if name in score_map:
                row[col] = score_map[name]
            else:
                row[col] = '未完成'
        else:
            row[col] = ''  # 树人学员不参加头狼课程，留空
    
    rows.append(row)

df_out = pd.DataFrame(rows)

# 计算树人学员平均分和头狼学员平均分
def calc_avg(row, course_cols, is_flag):
    """计算有效数字成绩的平均分"""
    if not is_flag:
        return ''
    scores = []
    for c in course_cols:
        v = row.get(c, '')
        if isinstance(v, (int, float)) and not pd.isna(v):
            scores.append(v)
        # 字符串'未完成'不计入
    if scores:
        return round(sum(scores) / len(scores), 1)
    return ''

sren_col_vals = sren_col_names
toul_col_vals = toul_col_names

# 添加完成率列
def calc_completion(row, course_cols, is_flag):
    if not is_flag:
        return ''
    completed = 0
    for c in course_cols:
        v = row.get(c, '')
        if isinstance(v, (int, float)) and not pd.isna(v):
            completed += 1
    total = len(course_cols)
    if total == 0:
        return ''
    return f'{completed}/{total}'

# 暂不加平均分列，直接输出原始成绩

print('输出列:', df_out.columns.tolist())
print(f'总行数: {len(df_out)}')

# ---- 8. 写出 Excel ----
output_path = 'C:/Users/Administrator/Desktop/树人计划第六期/学员成绩汇总.xlsx'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '成绩汇总'

# 定义样式
header_font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
header_fill_sren = PatternFill('solid', start_color='2E75B6')   # 蓝色-树人
header_fill_toul = PatternFill('solid', start_color='375623')   # 深绿-头狼
header_fill_base = PatternFill('solid', start_color='404040')   # 深灰-基础信息
cell_font = Font(name='微软雅黑', size=9)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 写表头
header_row = ['归属群', '序号', '姓名', '城市', '站点'] + sren_col_names + toul_col_names
for ci, h in enumerate(header_row, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border
    if ci <= 5:
        cell.fill = header_fill_base
    elif ci <= 5 + len(sren_col_names):
        cell.fill = header_fill_sren
    else:
        cell.fill = header_fill_toul

# 写数据
sren_member_fill = PatternFill('solid', start_color='DDEEFF')  # 浅蓝-树人成员行
toul_member_fill = PatternFill('solid', start_color='E8F0E0')  # 浅绿-头狼成员行
unfinished_fill = PatternFill('solid', start_color='FFE0E0')   # 浅红-未完成
score_font = Font(name='微软雅黑', size=9)
unfinished_font = Font(name='微软雅黑', size=9, color='CC0000')

all_col_keys = ['归属群', '序号', '姓名', '城市', '站点'] + sren_col_names + toul_col_names

# 按归属群排序：树人1群，头狼1群，头狼2群
df_out['_sort'] = df_out['归属群'].map({'树人1群': 0, '头狼1群': 1, '头狼2群': 2}).fillna(3)
df_out = df_out.sort_values(['_sort', '序号']).reset_index(drop=True)

for ri, row in df_out.iterrows():
    excel_row = ri + 2
    is_sren = '树人' in str(row.get('归属群', ''))
    row_bg = sren_member_fill if is_sren else toul_member_fill
    
    for ci, col in enumerate(all_col_keys, 1):
        val = row.get(col, '')
        cell = ws.cell(row=excel_row, column=ci, value=val)
        cell.font = score_font
        cell.alignment = center
        cell.border = border
        
        # 未完成单元格特殊标色
        if val == '未完成':
            cell.fill = unfinished_fill
            cell.font = unfinished_font
        else:
            cell.fill = row_bg

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 6
ws.column_dimensions['E'].width = 10
# 课程列宽度
for i in range(6, len(header_row) + 1):
    ws.column_dimensions[get_column_letter(i)].width = 14

# 冻结前5列和首行
ws.freeze_panes = 'F2'

# 第一行行高
ws.row_dimensions[1].height = 40

# ---- 9. 添加第二个sheet：说明 ----
ws2 = wb.create_sheet('折算说明')
ws2['A1'] = '折算规则说明'
ws2['A1'].font = Font(bold=True, size=12)
notes = [
    ['', ''],
    ['规则', '说明'],
    ['最高分保留', '同一学员在同一节课有多条成绩记录时，保留最高分'],
    ['折算100分制', '若该节课总分上限超过100分，则按比例折算：折算分 = 原始分 / 满分 * 100（保留1位小数）'],
    ['未完成', '学员在名单中但该节课无成绩记录，显示"未完成"'],
    ['空白', '该节课不属于该学员所在群组，不做统计'],
    ['', ''],
    ['树人成绩说明', ''],
]
for c in sren_courses:
    df = xl_sren[c] if c in xl_sren else None
    if df is not None and '总分' in df.columns:
        max_s = pd.to_numeric(df['总分'], errors='coerce').max()
        notes.append([c, f'满分={max_s}分' + ('（需折算）' if max_s > 100 else '（无需折算）')])
notes.append(['', ''])
notes.append(['头狼成绩说明', ''])
for c in toul_courses:
    df = xl_toul[c] if c in xl_toul else None
    if df is not None and '总分' in df.columns:
        max_s = pd.to_numeric(df['总分'], errors='coerce').max()
        col_label = toul_renamed[c]
        notes.append([col_label, f'满分={max_s}分' + ('（需折算）' if max_s > 100 else '（无需折算）')])

for r_data in notes:
    ws2.append(r_data)

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 30

wb.save(output_path)
print(f'\n已保存至: {output_path}')

# 打印匹配统计
print('\n=== 匹配统计 ===')
for c in sren_courses:
    score_map = sren_scores.get(c, {})
    matched = sum(1 for _, r in roster[roster['_is_sren']].iterrows() if r['姓名'] in score_map)
    total_sren = roster['_is_sren'].sum()
    print(f'[树人]{c}: {matched}/{total_sren} 人有成绩')

for c in toul_courses:
    score_map = toul_scores.get(c, {})
    matched = sum(1 for _, r in roster[roster['_is_toul']].iterrows() if r['姓名'] in score_map)
    total_toul = roster['_is_toul'].sum()
    print(f'[头狼]{c}: {matched}/{total_toul} 人有成绩')
